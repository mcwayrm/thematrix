from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook


WORKBOOK_PATH = Path("e:/matrix/thematrix/assets/naics_mapping_ai.xlsx")
BACKUP_SUFFIX = ".backup.xlsx"

COLS = {
    "prov": 4,
    "reg": 5,
    "nonmat": 6,
    "scope1": 7,
    "scope2": 8,
    "scope3": 9,
}

SCOPE2_ALL = ["EM 1.1", "EM 1.3", "EM 1.4"]

BUNDLES = {
    "office_reg": {
        "reg": ["SEEA 2.1", "SEEA 2.3", "SEEA 2.4", "SEEA 2.11", "SEEA 2.12", "SEEA 2.15"],
    },
    "industrial_reg": {
        "reg": ["SEEA 2.1", "SEEA 2.3", "SEEA 2.4", "SEEA 2.6", "SEEA 2.7", "SEEA 2.9", "SEEA 2.10", "SEEA 2.11", "SEEA 2.12", "SEEA 2.14", "SEEA 2.15"],
    },
    "ag_reg": {
        "reg": ["SEEA 2.1", "SEEA 2.2", "SEEA 2.5", "SEEA 2.6", "SEEA 2.9", "SEEA 2.10", "SEEA 2.11", "SEEA 2.12", "SEEA 2.17", "SEEA 2.18", "SEEA 2.19", "SEEA 2.20"],
    },
    "marine_reg": {
        "reg": ["SEEA 2.1", "SEEA 2.9", "SEEA 2.10", "SEEA 2.11", "SEEA 2.12", "SEEA 2.13", "SEEA 2.14", "SEEA 2.20"],
    },
    "crop": {
        "prov": ["SEEA 1.1", "SEEA 1.9"],
        "scope1": ["EM 1.1", "EM 1.4", "EM 1.7", "EM 1.11", "EM 2.9", "EM 2.20", "EM 2.21"],
        "scope3": ["EM 1.1", "EM 1.3", "EM 1.4", "EM 2.20", "EM 2.21", "EM 3.6"],
    },
    "animal": {
        "prov": ["SEEA 1.2", "SEEA 1.3", "SEEA 1.4", "SEEA 1.9"],
        "scope1": ["EM 1.1", "EM 1.3", "EM 1.4", "EM 1.11", "EM 2.18", "EM 2.19", "EM 2.20", "EM 2.21"],
        "scope3": ["EM 1.1", "EM 1.3", "EM 1.4", "EM 2.20", "EM 2.21", "EM 3.6"],
    },
    "forestry": {
        "prov": ["SEEA 1.5", "SEEA 1.7", "SEEA 1.8", "SEEA 1.9"],
        "scope1": ["EM 1.1", "EM 1.2", "EM 1.7", "EM 3.1"],
        "scope3": ["EM 1.1", "EM 1.2", "EM 1.3", "EM 3.1", "EM 3.6"],
    },
    "fishing": {
        "prov": ["SEEA 1.6", "SEEA 1.7", "SEEA 1.9"],
        "nonmat": ["SEEA 3.1", "SEEA 4.1"],
        "scope1": ["EM 1.1", "EM 1.3", "EM 1.7", "EM 1.8", "EM 1.9", "EM 3.1"],
        "scope3": ["EM 1.1", "EM 1.3", "EM 1.4", "EM 3.1", "EM 3.6"],
    },
    "extractive": {
        "prov": ["SEEA CF", "SEEA 1.9"],
        "scope1": ["EM 1.1", "EM 1.2", "EM 1.3", "EM 1.4", "EM 1.7", "EM 1.8", "EM 1.9", "EM 2.11", "EM 2.13", "EM 2.15", "EM 2.16", "EM 3.1", "EM 3.2", "EM 3.3", "EM 3.4", "EM 3.5"],
        "scope3": ["EM 1.1", "EM 1.3", "EM 1.4", "EM 3.1", "EM 3.2", "EM 3.6"],
    },
    "utility_power": {
        "prov": ["SEEA CF"],
        "scope1": ["EM 1.1", "EM 1.3", "EM 1.4", "EM 1.5", "EM 1.7", "EM 1.8", "EM 1.9"],
        "scope3": ["EM 1.1", "EM 1.3", "EM 1.4", "EM 3.1", "EM 3.6"],
    },
    "utility_water": {
        "prov": ["SEEA 1.9"],
        "scope1": ["EM 2.1", "EM 2.2", "EM 2.3", "EM 2.4", "EM 2.5", "EM 2.6", "EM 2.7", "EM 2.8", "EM 2.11", "EM 2.12", "EM 2.18", "EM 2.20", "EM 2.21"],
        "scope3": ["EM 1.1", "EM 1.3", "EM 1.4", "EM 2.11", "EM 3.6"],
    },
    "construction": {
        "prov": ["SEEA CF", "SEEA 1.9"],
        "reg": ["SEEA 2.3", "SEEA 2.4", "SEEA 2.6", "SEEA 2.12", "SEEA 2.14", "SEEA 2.15"],
        "nonmat": ["SEEA 3.2"],
        "scope1": ["EM 1.1", "EM 1.2", "EM 1.7", "EM 1.8", "EM 1.9", "EM 1.10", "EM 3.1", "EM 3.2", "EM 3.6"],
        "scope3": ["EM 1.1", "EM 1.3", "EM 1.4", "EM 3.1", "EM 3.6", "EM 3.7"],
    },
    "food_manuf": {
        "prov": ["SEEA 1.1", "SEEA 1.3", "SEEA 1.4", "SEEA 1.6", "SEEA 1.9"],
        "reg": ["SEEA 2.9", "SEEA 2.10", "SEEA 2.11", "SEEA 2.17", "SEEA 2.18", "SEEA 2.19"],
        "scope1": ["EM 1.1", "EM 1.3", "EM 1.4", "EM 1.7", "EM 1.10", "EM 1.11", "EM 2.6", "EM 2.7", "EM 2.8", "EM 2.18", "EM 2.20", "EM 2.21", "EM 3.6"],
        "scope3": ["EM 1.1", "EM 1.3", "EM 1.4", "EM 2.20", "EM 2.21", "EM 3.6", "EM 3.7"],
    },
    "bev_tobacco": {
        "prov": ["SEEA 1.1", "SEEA 1.9"],
        "reg": ["SEEA 2.9", "SEEA 2.10", "SEEA 2.11"],
        "scope1": ["EM 1.1", "EM 1.3", "EM 1.7", "EM 1.10", "EM 2.6", "EM 2.7", "EM 2.8", "EM 3.6"],
        "scope3": ["EM 1.1", "EM 1.3", "EM 1.4", "EM 3.6", "EM 3.7"],
    },
    "chemical": {
        "prov": ["SEEA CF", "SEEA 1.9"],
        "reg": ["SEEA 2.8", "SEEA 2.9", "SEEA 2.10", "SEEA 2.11"],
        "scope1": ["EM 1.1", "EM 1.3", "EM 1.5", "EM 1.6", "EM 1.7", "EM 1.8", "EM 1.9", "EM 1.10", "EM 2.3", "EM 2.8", "EM 2.10", "EM 2.11", "EM 2.12", "EM 2.13", "EM 2.14", "EM 2.16", "EM 3.1", "EM 3.2", "EM 3.7", "EM 3.9", "EM 3.10"],
        "scope3": ["EM 1.1", "EM 1.3", "EM 1.4", "EM 3.1", "EM 3.6", "EM 3.7", "EM 3.9", "EM 3.10"],
    },
    "heavy_manuf": {
        "prov": ["SEEA CF", "SEEA 1.9"],
        "scope1": ["EM 1.1", "EM 1.5", "EM 1.7", "EM 1.8", "EM 1.9", "EM 1.10", "EM 2.3", "EM 2.11", "EM 2.13", "EM 2.15", "EM 2.16", "EM 2.17", "EM 3.1", "EM 3.2", "EM 3.3", "EM 3.4", "EM 3.5", "EM 3.8", "EM 3.9"],
        "scope3": ["EM 1.1", "EM 1.3", "EM 1.4", "EM 3.1", "EM 3.6", "EM 3.7", "EM 3.8"],
    },
    "trade_general": {
        "reg": ["SEEA 2.1", "SEEA 2.3", "SEEA 2.4", "SEEA 2.11", "SEEA 2.12", "SEEA 2.15"],
        "scope1": ["EM 1.1", "EM 1.5", "EM 3.6", "EM 3.7"],
        "scope3": ["EM 1.1", "EM 1.3", "EM 1.4", "EM 3.6", "EM 3.7"],
    },
    "transport": {
        "reg": ["SEEA 2.1", "SEEA 2.3", "SEEA 2.4", "SEEA 2.11", "SEEA 2.12", "SEEA 2.14", "SEEA 2.15"],
        "scope1": ["EM 1.1", "EM 1.2", "EM 1.3", "EM 1.4", "EM 1.7", "EM 1.8", "EM 1.9", "EM 1.10", "EM 3.1"],
        "scope3": ["EM 1.1", "EM 1.3", "EM 1.4", "EM 3.1", "EM 3.6", "EM 3.7"],
    },
    "service_low": {
        "reg": ["SEEA 2.1", "SEEA 2.3", "SEEA 2.4", "SEEA 2.11", "SEEA 2.12", "SEEA 2.15"],
        "scope1": ["EM 1.1", "EM 1.5", "EM 3.6"],
        "scope3": ["EM 1.1", "EM 1.3", "EM 1.4", "EM 3.6"],
    },
    "arts_recreation": {
        "reg": ["SEEA 2.3", "SEEA 2.4", "SEEA 2.11", "SEEA 2.12", "SEEA 2.14", "SEEA 2.15"],
        "nonmat": ["SEEA 3.1", "SEEA 3.2", "SEEA 3.4", "SEEA 4.1"],
        "scope1": ["EM 1.1", "EM 1.5", "EM 3.6"],
        "scope3": ["EM 1.1", "EM 1.3", "EM 1.4", "EM 3.6"],
    },
    "education": {
        "reg": ["SEEA 2.3", "SEEA 2.4", "SEEA 2.11", "SEEA 2.12", "SEEA 2.15"],
        "nonmat": ["SEEA 3.3", "SEEA 4.1"],
        "scope1": ["EM 1.1", "EM 1.5", "EM 3.6"],
        "scope3": ["EM 1.1", "EM 1.3", "EM 1.4", "EM 3.6"],
    },
    "health": {
        "reg": ["SEEA 2.3", "SEEA 2.4", "SEEA 2.11", "SEEA 2.12", "SEEA 2.15"],
        "nonmat": ["SEEA 3.2", "SEEA 3.5"],
        "scope1": ["EM 1.1", "EM 1.5", "EM 2.18", "EM 2.19", "EM 3.6", "EM 3.10"],
        "scope3": ["EM 1.1", "EM 1.3", "EM 1.4", "EM 3.6", "EM 3.10"],
    },
    "public_admin": {
        "prov": ["SEEA 1.9"],
        "reg": ["SEEA 2.1", "SEEA 2.3", "SEEA 2.11", "SEEA 2.12", "SEEA 2.14", "SEEA 2.15"],
        "nonmat": ["SEEA 3.3", "SEEA 4.1"],
        "scope1": ["EM 1.1", "EM 1.5", "EM 3.6"],
        "scope3": ["EM 1.1", "EM 1.3", "EM 1.4", "EM 3.6"],
    },
}

PREFIX_BUNDLES = [
    ("11", ["ag_reg"]),
    ("111", ["crop"]),
    ("112", ["animal"]),
    ("113", ["forestry"]),
    ("114", ["fishing", "marine_reg"]),
    ("115", ["crop"]),
    ("21", ["extractive", "industrial_reg"]),
    ("2211", ["utility_power", "industrial_reg"]),
    ("2212", ["utility_power", "industrial_reg"]),
    ("2213", ["utility_water", "industrial_reg"]),
    ("23", ["construction"]),
    ("311", ["food_manuf"]),
    ("312", ["bev_tobacco"]),
    ("324", ["chemical", "industrial_reg"]),
    ("325", ["chemical", "industrial_reg"]),
    ("326", ["chemical", "industrial_reg"]),
    ("327", ["heavy_manuf", "industrial_reg"]),
    ("331", ["heavy_manuf", "industrial_reg"]),
    ("332", ["heavy_manuf", "industrial_reg"]),
    ("333", ["heavy_manuf", "industrial_reg"]),
    ("334", ["heavy_manuf", "industrial_reg"]),
    ("335", ["heavy_manuf", "industrial_reg"]),
    ("336", ["heavy_manuf", "industrial_reg"]),
    ("42", ["trade_general"]),
    ("44", ["trade_general"]),
    ("45", ["trade_general"]),
    ("48", ["transport"]),
    ("49", ["transport"]),
    ("51", ["service_low"]),
    ("52", ["service_low"]),
    ("53", ["service_low"]),
    ("54", ["education"]),
    ("55", ["service_low"]),
    ("56", ["service_low"]),
    ("61", ["education"]),
    ("62", ["health"]),
    ("71", ["arts_recreation"]),
    ("72", ["food_manuf", "arts_recreation"]),
    ("81", ["service_low"]),
    ("92", ["public_admin"]),
]

KEYWORD_BUNDLES = [
    (["aquaculture"], ["animal"]),
    (["fishing", "seafood"], ["fishing", "marine_reg"]),
    (["hunting", "trapping"], ["fishing"]),
    (["forest", "logging", "timber", "wood"], ["forestry"]),
    (["mining", "quarry", "oil", "gas extraction", "drilling", "coal", "ore"], ["extractive", "industrial_reg"]),
    (["water supply", "sewage", "sewer", "wastewater"], ["utility_water", "industrial_reg"]),
    (["electric power", "natural gas distribution", "steam"], ["utility_power", "industrial_reg"]),
    (["construction", "contractor", "highway", "bridge", "utility system"], ["construction"]),
    (["food manufacturing", "grocery", "dairy", "slaughter", "bakery", "tortilla", "fruit and vegetable preserving"], ["food_manuf"]),
    (["beverage", "tobacco"], ["bev_tobacco"]),
    (["petroleum", "chemical", "plastic", "rubber", "pharmaceutical", "medicine", "soap", "coating", "adhesive", "fertilizer", "pesticide"], ["chemical", "industrial_reg"]),
    (["metal", "machinery", "electronic", "electrical", "appliance", "transportation equipment", "semiconductor", "battery", "motor vehicle", "aerospace", "ship", "boat"], ["heavy_manuf", "industrial_reg"]),
    (["wholesale", "retail", "store"], ["trade_general"]),
    (["air transportation", "rail transportation", "water transportation", "truck transportation", "pipeline transportation", "courier", "messenger", "warehousing", "postal service"], ["transport"]),
    (["scenic", "sightseeing", "recreation", "hotel", "motel", "restaurant", "food services"], ["arts_recreation"]),
    (["data processing", "hosting", "software", "telecommunications", "information"], ["service_low"]),
    (["real estate", "rental"], ["service_low"]),
    (["education", "school", "college", "university", "scientific", "research"], ["education"]),
    (["hospital", "health care", "medical", "social assistance"], ["health"]),
    (["waste management", "remediation"], ["industrial_reg"]),
    (["arts", "entertainment", "museum", "park", "amusement"], ["arts_recreation"]),
    (["public administration"], ["public_admin"]),
]


def extend_unique(target, items):
    for item in items:
        if item not in target:
            target.append(item)


def apply_bundle(mapped, bundle_name):
    bundle = BUNDLES[bundle_name]
    for key, items in bundle.items():
        extend_unique(mapped[key], items)


def map_row(code, title):
    title_l = (title or "").lower()
    code_str = "".join(ch for ch in str(code) if ch.isdigit()) if code is not None else ""

    mapped = {key: [] for key in COLS}
    extend_unique(mapped["scope2"], SCOPE2_ALL)

    for prefix, bundle_names in PREFIX_BUNDLES:
        if code_str.startswith(prefix):
            for bundle_name in bundle_names:
                apply_bundle(mapped, bundle_name)

    for keywords, bundle_names in KEYWORD_BUNDLES:
        if any(keyword in title_l for keyword in keywords):
            for bundle_name in bundle_names:
                apply_bundle(mapped, bundle_name)

    return mapped


def main():
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Workbook not found: {WORKBOOK_PATH}")

    backup_path = WORKBOOK_PATH.with_name(WORKBOOK_PATH.stem + BACKUP_SUFFIX)
    copy2(WORKBOOK_PATH, backup_path)

    wb = load_workbook(WORKBOOK_PATH)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        code = ws.cell(row, 2).value
        title = ws.cell(row, 3).value
        if code is None or title is None:
            continue

        mapped = map_row(code, title)
        for key, col in COLS.items():
            ws.cell(row, col).value = "; ".join(mapped[key]) if mapped[key] else None

    wb.save(WORKBOOK_PATH)
    print({
        "saved": str(WORKBOOK_PATH),
        "backup": str(backup_path),
        "rows_processed": ws.max_row - 1,
    })


if __name__ == "__main__":
    main()
